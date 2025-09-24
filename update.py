import openpyxl
from openpyxl.styles import PatternFill

# 変更したい内容を辞書形式で定義します
# キー：セル番地, 値：新しい値
changes = {
    'B2': '更新されたテキスト',
    'C3': 456,
    'D4': '完了'
}

# ハイライト用の黄色い塗りつぶしスタイルを作成
yellow_fill = PatternFill(start_color='FFFFFF00',
                          end_color='FFFFFF00',
                          fill_type='solid')

input_file = 'original.xlsx'
output_file = 'updated.xlsx'

try:
    # 既存のExcelファイルを読み込む
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    # 定義した変更を一つずつ適用していく
    for cell_address, new_value in changes.items():
        # セルの値を更新
        sheet[cell_address].value = new_value
        # セルの背景色を黄色に設定
        sheet[cell_address].fill = yellow_fill
    
    # 変更を加えたファイルを新しい名前で保存
    workbook.save(output_file)
    print(f"'{input_file}' への変更を適用し、'{output_file}' として保存しました。")

except FileNotFoundError:
    print(f"エラー: ファイル '{input_file}' が見つかりません。")